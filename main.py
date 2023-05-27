from discord.ext import commands
import discord
import os
from dotenv import load_dotenv
from db import *
from LeagueHandler import *
import RiftChamps
from openpyxl import load_workbook
import random
from time import sleep

load_dotenv()

intents = discord.Intents.default()
intents.members = True
intents.message_content = True

bot = commands.Bot(command_prefix='$', intents=intents)

riftMatches = RiftChamps.Matches()


@bot.event
async def on_ready():
    print("we have logged in as {0.user}".format(bot))

#Max's Attempt
@bot.command(name="register")
async def register(message, IsMember, StudentID, Name): #Find a way to get the 3 arguments without asking for them from the user
    InsertUser(message.author.id, IsMember, StudentID, Name)
    await message.channel.send("registered")

@bot.command(name="setlolaccount")
async def setlolaccount(ctx, *args):
    Username = ""
    for arg in args:
         Username += (" " + arg)
    ID = UsernameToID(Username)
    if (ID == "Summoner name not valid"):
        await ctx.message.channel.send("Oops! Something went wrong. Did you input your username correctly?")
    else:
        EditUser(str(ctx.message.author.id), "RiotID", ID) #??? What RiotID and RiotName column?
        EditUser(str(ctx.message.author.id), "RiotName", Username)
        await ctx.message.channel.send("League of Legends account set to: "+ Username)

#@bot.command(name="RiftChampsStart")
async def RiftChampsLobbySetup(message):
    await message.channel.send("react to be added to the rift champs lobby!")

#@bot.command(name="RiftChampsMakeGames")
async def RiftChampsMakeGames(ctx, msgID):
    message = await ctx.message.channel.fetch_message(msgID)
    reactions = message.reactions
    IDlist = []
    for reaction in reactions:
        async for user in reaction.users():
            IDlist.append(user.id)
    
    Players = []
    for i in range(len(IDlist)):

        DiscordName = await bot.fetch_user(IDlist[i])  #wierd syntax maybe
        DiscordName = DiscordName.display_name
        RiotID = FetchRiotID(IDlist[i])
        RiotName = FetchRiotName(IDlist[i])
        
        Rank = GetRank(RiotID)
        print(Rank)
        RankValue = 1200
        if Rank[0] == "C":
            RankValue += 1300
        if Rank[0] == "G":
            RankValue += 1250
        if Rank[0] == "M":
            RankValue += 1200
        if Rank[0] == "D":
            RankValue += 1000
        if Rank[0] == "P":
            RankValue += 800
        if Rank[0] == "G":
            RankValue += 600
        if Rank[0] == "S":
            RankValue += 400
        if Rank[0] == "B":
            RankValue += 200
        if Rank[0] == "I":
            RankValue += 0
        
        if Rank[-1] == "1":
            RankValue += 150
        if Rank[-1] == "2":
            RankValue += 100
        if Rank[-1] == "3":
            RankValue += 50
        if Rank[-1] == "4":
            RankValue += 0



        # Player object holds all necessary information on each person that reacted
        Player = {"DiscordName" : DiscordName, "RiotName" : RiotName, "RankValue" : RankValue}
        Players.append(Player)
    GameList = RiftChamps.MakeGames(Players)
    MatchCount = 0
    while len(GameList) != 0:
        if len(GameList) >= 2:
            print(GameList)
            MatchCount += 1
            Team1 = GameList.pop(0)
            Team2 = GameList.pop(0)
            Team1String = "Team 1 Players:\n"
            Team2String = "Team 2 Players\n"
            
            for Player in Team1:
                Team1String += "\nDiscord: " + str(Player["DiscordName"]) + " - League of Legends:" + str(Player["RiotName"]) + " - Rank Value:" + str(Player["RankValue"])

            for Player in Team2:
                Team2String += "\nDiscord: " + str(Player["DiscordName"]) + " - League of Legends:" + str(Player["RiotName"]) + " - Rank Value:" + str(Player["RankValue"])

            embed=discord.Embed(
                title=("Match " + str(MatchCount)),
                url="https://www.youtube.com/watch?v=dQw4w9WgXcQ",
                description=(Team1String + "\n\n" + Team2String),
                color=0xFF5733)
            await ctx.send(embed=embed)
        
        else:
            print(GameList)
            MissedPlayers = GameList.pop(0)
            MissedPlayerString = "These players sadly missed out on a lobby:\n"

            for Player in MissedPlayers:
                MissedPlayerString += "\nDiscord: " + str(Player["DiscordName"]) + " - League of Legends:" + str(Player["RiotName"]) + " - Rank Value:" + str(Player["RankValue"])
        
            embed=discord.Embed(
                title=("Missed Players"),
                url="https://www.youtube.com/watch?v=dQw4w9WgXcQ",
                description=(MissedPlayerString),
                color=0xFF5733)
            await ctx.send(embed=embed)

#My attempt
@bot.command(name="RiftChampsSetup")
async def RiftChampsSetup(ctx):
    if committee(ctx.guild, ctx.author):
        await ctx.message.delete() #Need to figure out to link to a channel in a message, for #🤖bot-spam
        message = await ctx.channel.send('''League Rift Champs is beginning! React to this message to join.
        
Remember to register in #🤖bot-spam with "$AddSummoner <LeagueName>" while I'm online if you want to be a part of any Rift Champs event. Also remember to changes spaces to _ when entering your League name with $AddSummoner, like "UWA_Esports_Club_Bot".''')
        await message.add_reaction(await ctx.guild.fetch_emoji(781124340850229248))
        DM = ctx.author.dm_channel
        if (DM == None):
            DM = await ctx.author.create_dm()
        await DM.send('To begin Rift Champs, type "$RiftChampsStart {0}" in #{1}'.format(message.id, ctx.message.channel.name))
    else:
        raise Exception("User {0}#{1} tried $RiftChampsSetup".format(ctx.author.name, ctx.author.discriminator))

@bot.command(name="RiftChampsStart")
async def RiftChampsStart(ctx, id):
    await ctx.message.delete()
    if committee(ctx.guild, ctx.author):
        if str(id).isnumeric():
            try:
                setupMessage = await ctx.channel.fetch_message(int(id))
            except:
                await ctx.channel.send("An Error Occurred: id is incorrect")
                return None
            reaction = None
            for emoji in setupMessage.reactions:
                if not (type(emoji.emoji) is str):
                    if emoji.emoji.id == 781124340850229248: #TODO change this emoji id if the emoji changes
                        reaction = emoji
                        break
            if reaction == None:
                raise AttributeError("Reaction :LeagueOfLegends: not found")
            playersObjectRaw = [user async for user in reaction.users()]
            playersObject = [] #Everything from here till the end comment is to make committee go to the back of the list of people, so they're less likely to join lobbies
            committeeObject = []
            for user in playersObjectRaw:
                if committee(ctx.guild, user):
                    committeeObject.append(user)
                else:
                    playersObject.append(user)
            random.shuffle(playersObject)
            random.shuffle(committeeObject)
            playersObject.extend(committeeObject) #End of committee exclusion
            players = []
            nonMembers = []
            nonSignups = []
            requests = 0
            for player in playersObject:
                membership = Fetch(player.id, "IsMember")
                if membership != None:
                    if membership != 0:
                        if requests > 19: #Protects against exceeding the 20 requests per second restriction on the Riot API, got nothing against the 100 requests per 2 minutes restriction rn tho
                            requests = 0
                            sleep(1)
                        playerCurrentRank = GetPlayerRank(Fetch(player.id, "LoLID"))[0] #No protection against this returning an error string rn
                        requests += 1
                        EditUser(player.id, "Rank", playerCurrentRank)
                        players.append([player.id, Fetch(player.id, "LoLName"), playerCurrentRank])
                    else:
                        nonMembers.append([player.id, Fetch(player.id, "LoLName")])
                else:
                    nonSignups.append(player.id)
            await setupMessage.delete()
            leftOvers = riftMatches.createMatches(players)
            for match in riftMatches.matches:
                try: 
                    team2 = riftMatches.matches[match][1]
                except:
                    team = riftMatches.matches[match][0]
                    discord1 = ctx.guild.get_member(team[0][0]).mention
                    discord2 = ctx.guild.get_member(team[1][0]).mention
                    discord3 = ctx.guild.get_member(team[2][0]).mention
                    discord4 = ctx.guild.get_member(team[3][0]).mention
                    discord5 = ctx.guild.get_member(team[4][0]).mention
                    await ctx.channel.send('''***{Match}*** *(played in ARAMs and flat +5 points)*:
        **Team**: 
{Player1} - {Discord1}
{Player2} - {Discord2}
{Player3} - {Discord3}
{Player4} - {Discord4}
{Player5} - {Discord5}'''.format(Match = match, Player1 = team[0][1], Player2 = team[1][1], Player3 = team[2][1], Player4 = team[3][1], Player5 = team[4][1], Discord1 = discord1, Discord2 = discord2, Discord3 = discord3, Discord4 = discord4, Discord5 = discord5))
                else:
                    team1 = riftMatches.matches[match][0]
                    discord11 = ctx.guild.get_member(team1[0][0]).mention
                    discord12 = ctx.guild.get_member(team1[1][0]).mention
                    discord13 = ctx.guild.get_member(team1[2][0]).mention
                    discord14 = ctx.guild.get_member(team1[3][0]).mention
                    discord15 = ctx.guild.get_member(team1[4][0]).mention
                    discord21 = ctx.guild.get_member(team2[0][0]).mention
                    discord22 = ctx.guild.get_member(team2[1][0]).mention
                    discord23 = ctx.guild.get_member(team2[2][0]).mention
                    discord24 = ctx.guild.get_member(team2[3][0]).mention
                    discord25 = ctx.guild.get_member(team2[4][0]).mention
                    await ctx.channel.send('''***{Match}***:
        **Team1**: 
{Player11} - {Discord11}
{Player12} - {Discord12}
{Player13} - {Discord13}
{Player14} - {Discord14}
{Player15} - {Discord15}

        **Team2**:
{Player21} - {Discord21}
{Player22} - {Discord22}
{Player23} - {Discord23}
{Player24} - {Discord24}
{Player25} - {Discord25}'''.format(Match = match, Player11 = team1[0][1], Player12 = team1[1][1], Player13 = team1[2][1], Player14 = team1[3][1], Player15 = team1[4][1], Player21 = team2[0][1], Player22 = team2[1][1], Player23 = team2[2][1], Player24 = team2[3][1], Player25 = team2[4][1], Discord11 = discord11, Discord12 = discord12, Discord13 = discord13, Discord14 = discord14, Discord15 = discord15, Discord21 = discord21, Discord22 = discord22, Discord23 = discord23, Discord24 = discord24, Discord25 = discord25))
            if leftOvers:
                leftMsg = '''**People who missed out** *(Receive +5 points)*:'''
                for player in leftOvers:
                    EditUser(player[0], "Points", Fetch(player[0], "Points") + 5)
                    discordName = ctx.guild.get_member(player[0]).mention
                    leftMsg += '''
{Player} - {Discord}'''.format(Player = player[1], Discord = discordName)
                await ctx.channel.send(leftMsg)
            if nonMembers:
                notMsg = '''**People who aren't members but reacted** *(Receive no points)*:'''
                for player in nonMembers:
                    discordName = ctx.guild.get_member(player[0]).mention
                    notMsg += '''
{Player} - {Discord}'''.format(Player = player[1], Discord = discordName)
                await ctx.channel.send(notMsg)
            notMsg2 = '''**People who didn't signup but reacted** *(Receive no points)*:'''
            for playerID in nonSignups:
                if playerID != 818759928852250674:
                    discordName = ctx.guild.get_member(playerID).mention
                    notMsg2 += '''
{Discord}'''.format(Discord = discordName)
                if notMsg2 != '''**People who didn't signup but reacted** *(Receive no points)*:''':
                    await ctx.channel.send(notMsg2)
            DM = ctx.author.dm_channel
            if (DM == None):
                DM = await ctx.author.create_dm()
            await DM.send('''To close a match and award points, type "$MatchResult Match(x) (y)" anywhere on the server (preferably in the same channel tho). The (x) should be the number of the match, like Match1, and the (y) should be a number for which team won. Eg: "$MatchResult Match1 1" if Team1 won Match1. Same applies for 5-man match, type 1 if they won.

If you need to replace a player with another player, simply type "$ReplaceSummoner Match(x) <Discord name of person to be replaced> <Discord name of person who is replacing>" anywhere on the server. The (x) should be the number of the match.''')
        else:
            await ctx.channel.send("An Error Occurred: Message id is not numeric")
    else:
        raise Exception("User {0}#{1} tried $RiftChampsStart".format(ctx.author.name, ctx.author.discriminator))

@bot.command(name="MatchResult")
async def RiftChampsMatchOver(ctx, match, winner):
    await ctx.message.delete()
    if committee(ctx.guild, ctx.author):
        keys = [x for x in riftMatches.matches.keys()]
        if keys.count(match) != 0:
            finishedMatch = riftMatches.matches.pop(match)
            try:
                finishedMatch[1]
            except:
                discord1 = ctx.guild.get_member(finishedMatch[0][0][0]).display_name
                discord2 = ctx.guild.get_member(finishedMatch[0][1][0]).display_name
                discord3 = ctx.guild.get_member(finishedMatch[0][2][0]).display_name
                discord4 = ctx.guild.get_member(finishedMatch[0][3][0]).display_name
                discord5 = ctx.guild.get_member(finishedMatch[0][4][0]).display_name
                if int(winner) == 1:
                    for player in finishedMatch[0]:
                        EditUser(player[0], "Points", int(Fetch(player[0], "Points")) + 5)
                    await ctx.channel.send('''__{Match}__:
    **Winners**(ARAMs) *(Receive +5 points)*:
{Player1} - {Discord1}
{Player2} - {Discord2}
{Player3} - {Discord3}
{Player4} - {Discord4}
{Player5} - {Discord5}'''.format(Match = match, Player1 = finishedMatch[0][0][1], Player2 = finishedMatch[0][1][1], Player3 = finishedMatch[0][2][1], Player4 = finishedMatch[0][3][1], Player5 = finishedMatch[0][4][1], Discord1 = discord1, Discord2 = discord2, Discord3 = discord3, Discord4 = discord4, Discord5 = discord5))
                else:
                    for player in finishedMatch[0]:
                        EditUser(player[0], "Points", int(Fetch(player[0], "Points")) + 5)
                    await ctx.channel.send('''__{Match}__:
    **Losers**(ARAMs) *(Receive +5 points)*:
{Player1} - {Discord1}
{Player2} - {Discord2}
{Player3} - {Discord3}
{Player4} - {Discord4}
{Player5} - {Discord5}'''.format(Match = match, Player1 = finishedMatch[0][0][1], Player2 = finishedMatch[0][1][1], Player3 = finishedMatch[0][2][1], Player4 = finishedMatch[0][3][1], Player5 = finishedMatch[0][4][1], Discord1 = discord1, Discord2 = discord2, Discord3 = discord3, Discord4 = discord4, Discord5 = discord5))
            else:
                for player in finishedMatch[int(winner) - 1]:
                    EditUser(player[0], "Points", int(Fetch(player[0], "Points")) + 10)
                winners = finishedMatch.pop(int(winner) - 1)
                for player in finishedMatch[0]:
                    EditUser(player[0], "Points", int(Fetch(player[0], "Points")) + 5)
                discord11 = ctx.guild.get_member(winners[0][0]).display_name
                discord12 = ctx.guild.get_member(winners[1][0]).display_name
                discord13 = ctx.guild.get_member(winners[2][0]).display_name
                discord14 = ctx.guild.get_member(winners[3][0]).display_name
                discord15 = ctx.guild.get_member(winners[4][0]).display_name
                discord21 = ctx.guild.get_member(finishedMatch[0][0][0]).display_name
                discord22 = ctx.guild.get_member(finishedMatch[0][1][0]).display_name
                discord23 = ctx.guild.get_member(finishedMatch[0][2][0]).display_name
                discord24 = ctx.guild.get_member(finishedMatch[0][3][0]).display_name
                discord25 = ctx.guild.get_member(finishedMatch[0][4][0]).display_name
                await ctx.channel.send('''__{Match}__:
    **Winners** *(Receive +10 points)*:
{Player11} - {Discord11}
{Player12} - {Discord12}
{Player13} - {Discord13}
{Player14} - {Discord14}
{Player15} - {Discord15}

    **Losers** *(Receive +5 points)*:
{Player21} - {Discord21}
{Player22} - {Discord22}
{Player23} - {Discord23}
{Player24} - {Discord24}
{Player25} - {Discord25}'''.format(Match = match, Player11 = winners[0][1], Player12 = winners[1][1], Player13 = winners[2][1], Player14 = winners[3][1], Player15 = winners[4][1], Player21 = finishedMatch[0][0][1], Player22 = finishedMatch[0][1][1], Player23 = finishedMatch[0][2][1], Player24 = finishedMatch[0][3][1], Player25 = finishedMatch[0][4][1], Discord11 = discord11, Discord12 = discord12, Discord13 = discord13, Discord14 = discord14, Discord15 = discord15, Discord21 = discord21, Discord22 = discord22, Discord23 = discord23, Discord24 = discord24, Discord25 = discord25))
        else:
            await ctx.channel.send("An Error Occurred: There is no match with that name")
    else:
        raise Exception("User {0}#{1} tried $MatchResult".format(ctx.author.name, ctx.author.discriminator))
        
@bot.command(name="AddSummoner")
async def RiftChampsAdd(ctx, LoLName):
    user = ctx.author
    try:
        user.roles.index(ctx.guild.get_role(781114261480800296))
    except ValueError:
        isMember = 0
    else:
        isMember = 1
    LoLName = LoLName.replace("_", " ")
    id = GetPlayerID(LoLName)
    if type(id) is str:
        await ctx.channel.send("An Error Occurred: " + id)
    else:
        rank = GetPlayerRank(id)
        if type(rank) is str:
            await ctx.channel.send("An Error Occurred: " + rank)
        else:
            try:
                InsertUser(user.id, isMember, LoLName, id[0], rank[0])
            except:
                await ctx.channel.send("An Error Occurred: Trying to sign up when already signed up. Try '$UpdateSummoner Account <League Name>' to change account")
            else:
                await ctx.channel.send("Summoner {0} Added".format(LoLName))

@bot.command(name="UpdateSummoner")
async def RiftChampsUpdate(ctx, edit="Account", accountName = ""):
    error = RiftChampsEdit(ctx.author.id, edit, accountName)
    if type(error) == str:
        await ctx.channel.send("An Error Occurred: " + error)
    elif error == None:
        await ctx.channel.send("{0} updated!".format(edit.title()))

async def RiftChampsEdit(discordID, edit, accountName):
    if Fetch(discordID, "LoLID") != None:
        if edit.lower() == "name":
            EditUser(discordID, "LoLName", GetPlayerName(Fetch(discordID, "LoLID"))[0])
            return None
        elif edit.lower() == "account":
            if accountName != "":
                EditUser(discordID, "LoLName", accountName)
                id = GetPlayerID(accountName)
                if type(id) is str:
                    return id
                else:
                    EditUser(discordID, "LoLID", id[0])
                    rank = GetPlayerRank(id)
                    if type(rank) is str:
                        return rank
                    else:
                        EditUser(discordID, "Rank", rank[0])
                        return None
            else:
                return "Requires account name argument"
        return "Incorrect edit type"
    return "You're not registered"

@bot.command(name="ReplaceSummoner")
async def ReplaceSummoner(ctx, match, replaceeName, replacerName):
    riftMatch = riftMatches.get(match, None)
    if committee(ctx.guild, ctx.author):
        if riftMatch != None:
            replaceeDiscord = await ctx.guild.get_member_named(replaceeName)
            replacerDiscord = await ctx.guild.get_member_named(replacerName)
            if replaceeDiscord != None and replacerDiscord != None:
                replaceeRiot = Fetch(replaceeDiscord.id, "LoLName")
                replacerRiot = Fetch(replacerDiscord.id, "LoLName")
                if replaceeRiot != None and replacerRiot != None:
                    replacer = [replacerDiscord.id, replacerRiot, Fetch(replacerDiscord.id, "Rank")]
                    replacedMatch = RiftChamps.replace(riftMatch, replaceeRiot, replacer)
                    riftMatches.matches.update({match: replacedMatch})
                    await ctx.channel.send("Successfully replaced")
                else:
                    await ctx.channel.send("An Error Occurred: Either the replacee or replacer hasn't registered")
            else:
                await ctx.channel.send("An Error Occurred: Either the replacee or replacer name is wrong")
        else:
            await ctx.channel.send("An Error Occurred: There is no match with that name")

@bot.command(name="Leaderboard")
async def Leaderboard(ctx):
    await ctx.message.delete()
    if committee(ctx.guild, ctx.author):
        users = FetchAll(["DiscordID", "LoLName", "Points"])
        users.sort(reverse=True,key=points)
        message = "**Current Leaderboard**:"
        for user in users:
            message += '''
{Points:2} | {League} - {Discord}'''.format(Points=user[2], League=user[1], Discord=ctx.guild.get_member(user[0]).display_name)
        await ctx.channel.send(message)
    else:
        await ctx.channel.send("Currently a committee only command")

@bot.command(name="Commands")
async def PrintCommands(ctx):
    await ctx.message.delete()
    commandList = '''**Commands**:'''
    commandList += '''
    $AddSummoner (leagueName):\tTakes a league name as argument, collects the necessary data (league ID and Rank) from Riot servers, and then adds the information to a database that can be referred to for creating matches and awarding points. Need to replace spaces with underscores in your league name when registering. Will make a command to let you see your data in the future.
    $UpdateSummoner (type) (leagueName):\tUpdates your information in the database relevant to the type of update you entered. The types are (spelling is exact, case isn't):
        - Name: Updates your league name directly from Riot servers automatically
        - Account <leagueName>: Only type that uses your league name as an argument. Switches registered account to account with that name. Need to replace spaces with underscores in your league name when changing account.
    $Commands:\tLists all the commands you can use with the bot.'''
    DM = ctx.author.dm_channel
    if (DM == None):
        DM = await ctx.author.create_dm()
    await DM.send(commandList)
    if committee(ctx.guild, ctx.author):
        commandListCommittee = '''**Committee Commands**:'''
        commandListCommittee += '''
    $RiftChampsSetup:\tSends the reaction message for RiftChamps in the channel you are in. DMs you afterwards for the next command to continue RiftChamps.
    $RiftChampsStart (id):\tDeletes the setup message (given it's id as an argument), takes everyone that reacted and has registered, sorts them into matches which are sent as a message, and also sends a message containing everyone who missed out for various reasons. DMs you afterwards for the next command to close a match, and award points, for RiftChamps matches.
    $MatchResult (match) (teamNo):\tTakes the match name and a number representing the winning team as an argument, awards 10 points to the winners and 5 points to the losers, and then sends a message showing the winning and losing teams.
    $ReplaceSummoner (replacee) (replacer):\tTakes 2 discord names as arguments, the first is the replacee and the second is the replacer. Checks that everything is accounted for and then replaces the corresponding replacee player in the match with the replacer. Both players must be registered.
    $Leaderboard:\tSends a message containing every registered player's discord username, League username, and current points sorted from highest points to lowest.'''
        await DM.send(commandListCommittee)


#Below is a command to remove the member role from every person in the server and to assign the member role to select users
@bot.command(name="RefreshMemberRole") #Could do the thing where you iterate over the member list and give or take away member role based on if their username is in the excel sheet (currently other way round)
async def RefreshMemberRole(ctx, type="Assign"):      #Could also have it to if I do the thing above, that both lists are sorted alphabetically, and delete from the excel list as we go, so if the current member isn't in the first ten spots on the list then they don't exist. This isn't logically consistent but it's an idea
    server = ctx.guild                 #Would have to rejoin clearRole and assignRoleFromExcel to make the above happen so not going to do that now
    if committee(server, ctx.author):
        role = server.get_role(781114261480800296) #TODO If new role is used or role ids change, change this role ID to match
        if type == "Both":
            await clearRole(server, role)
            await assignRoleFromExcel(server, role)
        elif type == "Remove":
            await clearRole(server, role)
        elif type == "Assign":
            await assignRoleFromExcel(server, role)

async def clearRole(server, role): #Removes a role from every member on a server/guild
    for user in server.members:
        try: #This exception catcher is just to check if the member even has the role so that we don't waste time performing an await function on every member
            user.roles.index(role)
        except ValueError:
            pass
        else:
            await user.remove_roles(role)

async def assignRoleFromExcel(server, role): #Gives a role to members on a server/guild based on if their name shows up in an excel table linked in the next line of code
    workbook = load_workbook(filename=os.getenv("MEMBER_LIST_PATH")) #TODO Change file directory to actual excel file directory before use, also make sure it's a file path/directory and not just the file name
    sheet = workbook["exportedtable_18_05_2023"] #TODO Change sheet name to actual sheet name in excel file before use
    nameColumn = "!"
    for i in range(26): #Could find a better way to do this but essentially this loop just takes i through the letters of the alphabet. If the column is on cell AA1, well fuck
        if sheet["{0}1".format(chr(i+65))].value == "What is your Discord name and id? (e.g Swole Joel#5279)": #TODO Change column name to actual column name for discord names before use
            nameColumn = chr(i+65)
            break
    if nameColumn != "!":
        unassigned_names = []
        duplicate_names = []
        memberList = list(server.members).copy() #Copies the list of members on the server so that I can sort it alphabetically by name and perform a binary search for a larger server
        memberList.sort(key=getDiscordName)
        for i in range(2, sheet.max_row+1): #Lists through the values under the Discord Name column
            currentName = sheet["{0}{1}".format(nameColumn, i)].value
            if currentName != None:
                found = False
                start = 0 #Beginning from here is just an exceptionally long binary search through the list of server members, but it should be faster than linear searching if the server's got a lot of members
                end = len(memberList) - 1
                while start <= end:
                    pointer = (start + end) // 2
                    username = getDiscordName(memberList[pointer])
                    if currentName < username:
                        end = pointer - 1
                    elif currentName > username:
                        start = pointer + 1
                    else:
                        found = True
                        try: #Super long and probably really unoptimised way around causing an index error for going above or below a legal index for the member list
                            memberList[pointer + 1]
                        except IndexError:
                            if getDiscordName(memberList[pointer - 1]) != username: #This and similar if statements are essentially to make sure there aren't any duplicate discord names on the server
                                try:                                          #And this and similar exception catchers are a way of checking if someone already has the role so that we don't have to waste time giving it to them again
                                    memberList[pointer].roles.index(role)
                                except ValueError:
                                    await memberList[pointer].add_roles(role)
                            else:
                                duplicate_names.append(currentName)
                        else:
                            try:
                                memberList[pointer - 1]
                            except IndexError:
                                if getDiscordName(memberList[pointer + 1]) != username:
                                    try: 
                                        memberList[pointer].roles.index(role)
                                    except ValueError:
                                        await memberList[pointer].add_roles(role)
                                else:
                                    duplicate_names.append(currentName)
                            else:
                                if getDiscordName(memberList[pointer + 1]) != username and getDiscordName(memberList[pointer - 1]) != username:
                                    try: 
                                        memberList[pointer].roles.index(role)
                                    except ValueError:
                                        await memberList[pointer].add_roles(role)
                                else:
                                    duplicate_names.append(currentName)
                        break
                if not found:
                    unassigned_names.append(currentName)
        print("The unassigned names are " + str(unassigned_names)) #Prints to the terminal any unassigned members from the excel sheet
        print("And the duplicate names are " + str(duplicate_names)) #Prints to the terminal any duplicate names on the server that match a name on the excel sheet

def getDiscordName(user): #Returns the discord name of a discord member object
    return user.name + "#" + user.discriminator

def points(row):
    return row[2]

def committee(server, user): #Passes a server/guild and user to see if the user has the committee role in the server/guild, returns True if they do and vise versa
    try:
        user.roles.index(server.get_role(781114316157747220)) #TODO if the committee role is changed to a new role, change this to match
    except ValueError:
        return False
    else:
        return True





bot.run(os.getenv("DISCORD_TOKEN"))

#sampletable = [[1, "SantyMax", 2000],[2, "SantyMax2", 2250],[3, "SantyMax3", 2000],[4, "SantyMax4", 1600],[5, "SantyMax5", 1400],[6, "SantyMax6", 1200],[7, "SantyMax7", 2500],[8, "SantyMax8", 1600],[9, "SantyMax9", 2250],[10, "SantyMax10", 2450]]